"""Business logic for the Budget Checker.

This module loads an Excel workbook from a remote URL (configured via
``BUDGET_SOURCE_URL`` environment variable) and exposes functions to build
dropdown selections and perform budget checks. Data is cached for a
configurable TTL (via ``CACHE_TTL`` environment variable) to avoid
re-downloading the workbook on every request.

The structure of the workbook is expected to match the legacy
``Budget cheeker.py`` script: the ``AllSubBAs`` sheet contains rows with
Year in column A, BA# in column C, Contract Type in column D, Project Name
in column K, Supplier in column L, Remaining Budget in column AH and
Status in column AN. Only rows from the year 2026 are considered.
"""

from __future__ import annotations

import os
import threading
import time
import tempfile
import shutil
from typing import List, Dict, Any, Optional

import requests
import openpyxl

__all__ = ["get_budget_dropdowns", "check_budget"]

# Configured via environment variables
BUDGET_SOURCE_URL = os.getenv("BUDGET_SOURCE_URL")
CACHE_TTL = int(os.getenv("CACHE_TTL", "300"))  # seconds

if not BUDGET_SOURCE_URL:
    raise RuntimeError(
        "BUDGET_SOURCE_URL environment variable is required for budget checker."
    )

_cache_lock: threading.Lock = threading.Lock()
_cached_rows: Optional[List[tuple]] = None
_last_load_ts: float = 0.0


def _download_workbook(url: str) -> str:
    """Download the Excel file from a URL into a temporary file and return its path."""
    response = requests.get(url, stream=True)
    response.raise_for_status()
    # Create temp file on disk. Using delete=False because we'll remove manually.
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    with os.fdopen(fd, "wb") as f:
        for chunk in response.iter_content(chunk_size=8192):
            if chunk:
                f.write(chunk)
    return path


def _load_rows() -> List[tuple]:
    """Load the data rows from the Excel workbook. Uses read_only mode for efficiency."""
    tmp_path = _download_workbook(BUDGET_SOURCE_URL)
    try:
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        if "AllSubBAs" not in wb.sheetnames:
            raise RuntimeError(
                "'AllSubBAs' sheet not found in budget workbook."
            )
        rows = list(wb["AllSubBAs"].iter_rows(min_row=2, values_only=True))
        wb.close()
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
    return rows


def _get_rows() -> List[tuple]:
    """Get cached rows, reloading if the cache has expired."""
    global _cached_rows, _last_load_ts
    now = time.monotonic()
    with _cache_lock:
        if _cached_rows is None or (now - _last_load_ts) > CACHE_TTL:
            _cached_rows = _load_rows()
            _last_load_ts = now
    return _cached_rows  # type: ignore[return-value]


def get_budget_dropdowns() -> Dict[str, List[str]]:
    """Extract unique values for the Contract Type, Project Name and Supplier dropdowns.

    Returns a dict with keys ``box1``, ``box2``, ``box3``.
    """
    rows = _get_rows()
    box1: set[str] = set()
    box2: set[str] = set()
    box3: set[str] = set()

    for row in rows:
        # Only consider rows from year 2026; Year is column A (index 0)
        try:
            year_val = int(row[0]) if row[0] is not None else None
        except (TypeError, ValueError):
            continue
        if year_val != 2026:
            continue
        # Column D (index 3): Contract Type
        ct = row[3]
        # Column K (index 10): Project Name
        pn = row[10]
        # Column L (index 11): Supplier
        sp = row[11]
        if ct is not None:
            box1.add(str(ct).strip())
        if pn is not None:
            box2.add(str(pn).strip())
        if sp is not None:
            box3.add(str(sp).strip())

    return {
        "box1": sorted(box1),
        "box2": sorted(box2),
        "box3": sorted(box3),
    }


def check_budget(contract_type: str, project_name: str, supplier: str, requested_amount: float) -> Dict[str, Any]:
    """Perform the budget check for the given inputs.

    Returns a dict containing either ``results`` (list of row results) or ``error`` with details.

    Each result item includes:
        - ba_number: string (BA#)
        - remaining_budget: float
        - requested: float (input value)
        - gap: float (requested minus remaining)
        - budget_status: 'over_budget' or 'proceed'
        - sheet_status: string from column AN (Status)
    """
    if not contract_type or not project_name or not supplier:
        return {
            "error": "Contract Type, Project Name and Supplier are required."
        }
    try:
        requested_amount = float(requested_amount)
    except (TypeError, ValueError):
        return {
            "error": "Requested amount must be numeric."
        }

    rows = _get_rows()
    results: List[Dict[str, Any]] = []
    skipped_count = 0
    # Column index map (0-based): A=0, C=2, D=3, K=10, L=11, AH=33, AN=39
    for idx, row in enumerate(rows):
        if len(row) < 40:
            continue
        # Year must be 2026
        try:
            year_val = int(row[0]) if row[0] is not None else None
        except (TypeError, ValueError):
            continue
        if year_val != 2026:
            continue
        # Match Contract Type, Project Name, Supplier
        row_contract = str(row[3]).strip() if row[3] is not None else ""
        row_project = str(row[10]).strip() if row[10] is not None else ""
        row_supplier = str(row[11]).strip() if row[11] is not None else ""
        if row_contract != contract_type or row_project != project_name or row_supplier != supplier:
            continue
        # Extract numeric fields
        try:
            ba_number = row[2]
            remaining_budget = float(row[33])
        except (TypeError, ValueError):
            skipped_count += 1
            continue
        sheet_status = (
            str(row[39]).strip() if row[39] is not None else "—"
        )
        gap = requested_amount - remaining_budget
        budget_status = "over_budget" if gap > 0 else "proceed"
        results.append(
            {
                "ba_number": str(ba_number) if ba_number is not None else "—",
                "remaining_budget": remaining_budget,
                "requested": requested_amount,
                "gap": gap,
                "budget_status": budget_status,
                "sheet_status": sheet_status,
            }
        )

    if not results:
        detail = (
            f" (skipped {skipped_count} rows with bad data)" if skipped_count else ""
        )
        return {
            "error": (
                f'No matching record found (Year 2026) for Contract Type "{contract_type}", '
                f'Project Name "{project_name}", Supplier "{supplier}"{detail}.'
            )
        }
    return {"results": results}